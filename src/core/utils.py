import requests
from astroquery.gaia import Gaia
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import time
import logging
import pandas as pd
import numpy as np

#-------------------------------------------------------------------------------------------------- 

def execute_gaia_query(query, str_columns=None, output_file=None, retries=3, delay=5):
    """
    Executes a Gaia query and optionally saves the results to an Excel file.

    Parameters:
    -----------
    query : str
        The ADQL query to execute.
    str_columns : list, optional
        List of column names to convert to string type.
    output_file : str, optional
        Path to save the Excel file. If None, no file is saved.
    retries : int
        Number of times to retry the query in case of failure.
    delay : int
        Delay in seconds between retries, with exponential backoff.

    Returns:
    --------
    pandas.DataFrame or None
        Query results as a DataFrame, or None if the query fails.
    """
    
    # Suppress the specific info message from astroquery
    logging.getLogger('astroquery').setLevel(logging.ERROR)

    attempt = 0
    while attempt < retries:
        try:
            # Execute the query
            job = Gaia.launch_job_async(query)
            df = job.get_results().to_pandas()

            # Convert specified columns to string
            if str_columns:
                for col in str_columns:
                    if col in df.columns:
                        df[col] = df[col].astype(str)

            # Save to Excel if a filename is provided
            if output_file:
                df.to_excel(output_file, index=False)
                adjust_column_widths(output_file)
                print(f"{len(df)} stars retrieved")
                print(f"Query results saved to {output_file}")
            return df

        except requests.exceptions.HTTPError as e:
            print(f"An HTTP error occurred: {e}")
        except requests.exceptions.ConnectionError as e:
            print(f"A connection error occurred: {e}")
        except Exception as e:
            print(f"An error occurred: {e}")

        attempt += 1
        print(f"Retrying in {delay} seconds... (Attempt {attempt}/{retries})")
        time.sleep(delay)
        delay *= 2  # Exponential backoff

    print("Failed to execute query after several attempts.")
    return None

#--------------------------------------------------------------------------------------------------

def adjust_column_widths(excel_file, highlight_columns=None):
    """
    Adjusts the column widths of an Excel workbook based on the maximum length of data in each column,
    freezes the first row, and optionally highlights specific columns.

    Parameters:
    - excel_file: str, the path to the Excel file to be processed.
    - highlight_columns: list, optional, a list of column names to highlight in the Excel workbook.
    """
    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file)
    worksheet = workbook.active

    # Adjust the column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Get the column letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        worksheet.column_dimensions[column_letter].width = max_length + 2

    # Freeze the first row
    worksheet.freeze_panes = worksheet['A2']

    # Highlight specific columns if provided
    if highlight_columns:
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
        bold_font = Font(bold=True)

        # Get the column headers from the first row
        headers = {cell.value: cell.column for cell in worksheet[1]}  # Map header name to column index
        for col_name in highlight_columns:
            if col_name in headers:
                col_idx = headers[col_name]  # Get the column index
                for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=worksheet.max_row):
                    for c in cell:
                        c.fill = highlight_fill
                        if c.row == 1:  # Apply bold font to the header row
                            c.font = bold_font

    # Save the workbook
    workbook.save(excel_file)

#--------------------------------------------------------------------------------------------------

def save_and_adjust_column_widths(df, output_file, highlight_columns=None):
    """
    Saves a DataFrame to an Excel file, adjusts column widths, freezes the first row,
    and optionally highlights specific columns.

    Parameters:
    - df: pd.DataFrame, the DataFrame to save.
    - output_file: str, the path to the output Excel file.
    - highlight_columns: list, optional, a list of column names to highlight in the Excel workbook.
    """
    df.to_excel(output_file, index=False)
    adjust_column_widths(output_file, highlight_columns=highlight_columns)
    print(f"Results saved to {output_file}")
#--------------------------------------------------------------------------------------------------

def angular_separation_vectorized(ra1, dec1, ra2_array, dec2_array):
    """Vectorized calculation of angular separation"""
    ra1, dec1 = np.radians([ra1, dec1])
    ra2_array, dec2_array = np.radians([ra2_array, dec2_array])
    
    delta_ra = ra2_array - ra1
    delta_dec = dec2_array - dec1
    
    a = np.sin(delta_dec/2)**2 + np.cos(dec1) * np.cos(dec2_array) * np.sin(delta_ra/2)**2
    c = 2 * np.arcsin(np.sqrt(a))
    
    return np.degrees(c)

#-------------------------------------------------------------------------------------------------- 

def format_gaia_id(gaia_id):
    """Convert GAIA ID to string format without scientific notation"""
    if pd.isna(gaia_id):
        return None
    gaia_id_str = str(gaia_id).replace('Gaia DR2 ', '')
    if 'e' in gaia_id_str.lower():
        return f"{float(gaia_id_str):.0f}"
    return gaia_id_str